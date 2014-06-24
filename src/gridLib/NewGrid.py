# -*- coding:utf-8 -*-
"""
Created on 09/12/2010

@author: usuario
"""
__all__ = ['NewGrid']

DEFAULT_GRID_SIZE= (0,0)
DEFAULT_FONT_SIZE = 12
DECIMAL_POINT = '.'

## THIRD PART LIBRARIES
import os
import functools

import wx
import wx.grid
from numpy import ndarray, ravel

import xlrd
from openpyxl import Workbook, load_workbook


# OWN LIBRARIES
from imagenes import imageEmbed
from slbTools import ReportaExcel
from easyDialog import Dialog as dialog
from easyDialog import Ctrl, Busy
from slbTools import isnumeric, isiterable
from .GridCopyPaste import PyWXGridEditMixin, MyContextGrid
from .gridCellRenderers import floatRenderer
from gridLib import gridEditors as GE

COLTYPES= {'0':'DATE','1':'LIST','2':'FLOAT','3':'VARCHAR','4':'INTEGER'}

###########################################################################
## Class NewGrid
###########################################################################
class NewGrid(wx.grid.Grid, PyWXGridEditMixin):
    def __init__(self, parent, size= (1,2), *args, **params):
        wx.grid.Grid.__init__(self, parent, *args, **params)
        PyWXGridEditMixin.__init__(self)
        try:  self.__name= params.pop('name')
        except: self.__name = 'selobu'
        for key, value in params.items():
            if not key.startswith('_'):
                setattr(self,key, value)
        self._sh= None
        self.path2file= None
        self.maxrow=   0
        self.maxcol=   0
        self.__zoom=   1.0
        self.moveTo=   None
        self.wildcard= __("Suported Formats")+" (*.xls;*xlsx;*.txt;*csv)|*.xls;*xlsx;*.txt;*csv" + \
                       __("All Files")+" (*.*)|*.*"
        self.usedCols= ([], [],)
        # </p>
        
        #################################
        # Grid
        self.CreateGrid( size[0], size[1] )

        self.__columnTypes= [3]*size[1]
        self.EnableEditing( True )
        self.EnableGridLines( True )
        self.EnableDragGridSize( False )
        self.SetMargins( 0, 0 )
        self.floatCellAttr= None
        # Columns
        self.EnableDragColMove( False )
        self.EnableDragColSize( True )
        self.SetColLabelSize( 30 )
        self.SetColLabelAlignment( wx.ALIGN_CENTRE, wx.ALIGN_CENTRE )
        # Rows
        self.EnableDragRowSize( True )
        self.SetRowLabelSize( 80 )
        self.SetRowLabelAlignment( wx.ALIGN_CENTRE, wx.ALIGN_CENTRE )
        try:          self.defaultRowSize=  self.GetRowSize(0)
        except:       self.defaultRowSize=  19
        try:          self.defaultColSize=  self.GetColSize(0)
        except:       self.defaultColSize=  80
        self.defaultRowLabelSize= self.GetRowLabelSize()
        self.defaultColLabelSize= self.GetColLabelSize()
        # Label Appearance
        if wx.Platform == '__WXMAC__':
            ##self.SetGridLineColour("#b7b7b7")
            self.SetGridLineColour( wx.BLACK)
            self.SetLabelBackgroundColour("#d2d2d2")
            self.SetLabelTextColour("#444444")
        else:
            self.SetLabelBackgroundColour( wx.Colour( 254, 226, 188 ) )
        # Cell Defaults
        self.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
        self.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        #################################

        self.Bind(wx.grid.EVT_GRID_EDITOR_CREATED,         self.onCellEdit)
        self.Bind(wx.grid.EVT_GRID_CMD_LABEL_RIGHT_DCLICK, self.RangeSelected)
        self.Bind(wx.grid.EVT_GRID_CELL_CHANGED,           self.onCellChangedIncreaseSize)
        self.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK,       self.OnGridRighClic)
        self.Bind(wx.grid.EVT_GRID_LABEL_LEFT_DCLICK,      self.__onGridCmdLabelLeftDClick)
        self.Bind(wx.EVT_MOUSEWHEEL,                       self.__OnMouseWheel)
        #self.Bind(EVT_GRID_PASTE,                          self.onPaste)

    def __controlColnumber(func):
        """
        a decorator to control and transform the input column from string or number into a valid number
        """
        def getColRefAsNumber(self,colNumber):
            """given a number or a name of a column return its equivalent number, planed to be used as a decorator"""
            if isinstance( colNumber, (str, unicode)):
                colN= colNumber[:]
                if not(colN in self.colNames):
                    raise TypeError('You can only use a numeric value, or the name of an existing column')
                for pos, value in enumerate(self.colNames):
                    if value == colN:
                        colN= pos
                        break
            else:
                colN = colNumber
            colN= int(colN)
            if colN < 0 or colN > self.NumberCols:
                raise StandardError( __('The minimum accepted col is 0, and the maximum is %i')%self.GetNumberCols()-1)
            return colN

        @functools.wraps(func)
        def wrapper(self,*args, **params):
            args = list(args)
            args[0] = getColRefAsNumber(self, args[0]) # the first argument is self
            args = tuple(args)
            return func(self,*args, **params)
        return wrapper

    ###############  PROPERTIES #############
    @property
    def zoom(self):
        return self.__zoom
    @zoom.setter
    def zoom(self, zoom):
        if isnumeric(zoom):
            if zoom > 0 and zoom < 10:
                self.__zoom = zoom
    @property
    def availableColTypes(self):
        return range(4)
    @property
    def name(self):
        return self.__name
    @name.setter
    def name(self, name):
        self.__name = name
    @property
    def rowNames( self):
        return [self.GetRowLabelValue(row) for row in range(self.GetNumberRows())]
    @rowNames.setter
    def rowNames( self, rowNames):
        if isinstance(rowNames, (str, unicode)):
            rolNames= [rowNames]

        if not isiterable(rowNames):
            raise TypeError('rowNames must be an iterable object')

        if len(rowNames) == 0:
            return

        for rownumber, rowname in enumerate(rowNames):
            self.SetRowLabelValue(rownumber, rowname)
    @property
    def colNames(self):
        return [self.GetColLabelValue(col) for col in range(self.GetNumberCols())]
    @colNames.setter
    def colNames(self, colNames):
        if isinstance(colNames, (str, unicode)):
            colNames= [colNames]

        if not isiterable(colNames):
            raise TypeError('colNames must be an iterable object')

        if len(colNames) == 0:
            return

        for colnumber, colname in enumerate(colNames):
            self.SetColLabelValue(colnumber, colname)
    ###############################################     
    #
    def __onGridCmdLabelLeftDClick(self, evt):
        """"TO CHANGING THE TYPE OF A COLUMN"""
        # identifying the column
        columnNumber= evt.GetCol()
        colNames= self.colNames[:]
        currName= colNames[columnNumber]
        # show a dialog to change the name
        dlg= dialog()
        txt1=  Ctrl.StaticText( __('Column name'))
        txt2=  Ctrl.StaticText( __('Column type'))
        edit1= Ctrl.TextCtrl(currName)
        currColType= self._getColType(columnNumber)
        edit2= Ctrl.Choice(['DATE','LIST','FLOAT','VARCHAR','INTEGER'],currColType)
        dlg.struct= [[txt1, edit1 ],
                     [txt2, edit2 ]]
        if dlg.ShowModal() == wx.ID_OK:
            values= dlg.GetValue()
            dlg.Destroy()
        else:
            dlg.Destroy()
            return
        newColName,colType = values
        if newColName in ('',u''):
            print __("You has been input an invalid column Name")
            return
        colNames[columnNumber] = u''
        if newColName in colNames:
            print __("The new name of the column already exist!")
            return
        self.setColType(columnNumber, colType )
        colNames[columnNumber] = newColName
        # updating the name of the column
        self.colNames= colNames
        evt.Skip()
    def _getColType(self, colNumber):
        # adjusting the size of the column tipes
        # because of the ausence of the addcolumn evt
        diff = self.GetNumberCols() - len(self.__columnTypes)
        if diff > 0:
            self.__columnTypes.extend([3]*diff) # 3 is used to set the colour of the rows
        elif diff < 0:
            for i in range(-diff):
                self.__columnTypes.pop(-1)
        return self.__columnTypes[colNumber]

    @__controlColnumber
    def getColumnTypeAsStr(self, colNumber):
        return COLTYPES[colNumber.__str__()]

    def _selectDbTableDialog_initAttr(self):
        renderer = floatRenderer( 4)
        self.__oddAttr=  wx.grid.GridCellAttr()
        self.__oddAttr.SetBackgroundColour( wx.Colour( 220, 230, 250 ))
        self.__oddAttr.SetRenderer( renderer)
        self.__evenAttr= wx.grid.GridCellAttr()
        self.__evenAttr.SetBackgroundColour( wx.Colour( 146, 190, 183 ))
        self.__evenAttr.SetRenderer( renderer)

    def GetAttr(self, row, col, kind):
        attr=   [ self.__evenAttr, self.__oddAttr][row % 2]
        ##attr=  attr[col] # renderer by col is missing
        attr.IncRef()
        return attr

    def __zoom_rows(self):
        """Zooms grid rows"""
        for rowno in xrange(self.GetNumberRows()):
            self.SetRowSize(rowno, self.defaultRowSize*self.zoom)
        self.SetRowLabelSize( self.defaultRowLabelSize * self.zoom)

    def __zoom_cols(self):
        """Zooms grid columns"""
        tabno = 1 #self.current_table
        for colno in xrange(self.GetNumberCols()):
            self.SetColSize(colno, self.defaultColSize * self.zoom)
        self.SetColLabelSize(self.defaultColLabelSize * self.zoom)

    def __zoom_labels(self):
        """Zooms grid labels"""
        labelfont = self.GetLabelFont()
        labelfont.SetPointSize(max(1, int(DEFAULT_FONT_SIZE * self.zoom)))
        self.SetLabelFont(labelfont)

    ###############################
    ### EVENTS
    def __OnMouseWheel(self, event):
        """Event handler for mouse wheel actions
        Invokes zoom when mouse when Ctrl is also pressed
        """
        if event.ControlDown():
            zoomstep = 0.05 * event.LinesPerAction
            if event.WheelRotation > 0:
                self.zoom += zoomstep
            else:
                if self.zoom > 0.6:
                    self.zoom -= zoomstep
            self.__zoom_rows()
            self.__zoom_cols()
            self.__zoom_labels()
            self.ForceRefresh()
        else:
            event.Skip()
    #def onCellChanged(self, evt):
    #    self.hasChanged= True
    #    self.hasSaved=   False
    #    col, row=  evt.GetCol(), evt.GetRow()
    #    self.__cellChanged(row,col)
    #    evt.Skip()

    def onCellChangedIncreaseSize(self, evt):
        """to increase the size of the grid by one row or by one column"""
        numCols, numRows = self.GetNumberCols(), self.GetNumberRows()
        currColNum, currRowNum = evt.Col, evt.Row
        if currColNum == numCols -1:
            self.AppendCols(1)
            self.SetGridCursor(currRowNum, numCols)
        if currRowNum == numRows-1:
            self.AppendRows(1)
            self.SetGridCursor(numRows, currColNum)
        # refresh the sheet
        #self.ForceRefresh()
        evt.Skip()
    def onCellEdit(self, event):
        '''
        When cell is edited, get a handle on the editor widget
        and bind it to EVT_KEY_DOWN
        '''        
        editor = event.GetControl()        
        editor.Bind(wx.EVT_KEY_DOWN, self.onEditorKey)
        event.Skip()

    def onEditorKey(self, event):
        '''
        Handler for the wx's cell editor widget's keystrokes. Checks for specific
        keystrokes, such as arrow up or arrow down, and responds accordingly. Allows
        all other key strokes to pass through the handler.
        '''
        keycode = event.GetKeyCode() 
        if keycode == wx.WXK_UP:
            self.MoveCursorUp(False)
        elif keycode == wx.WXK_DOWN:
            self.MoveCursorDown(False)
        elif keycode == wx.WXK_LEFT:
            self.MoveCursorLeft(False)
        elif keycode == wx.WXK_RIGHT:
            self.MoveCursorRight(False)
        else:
            pass
        event.Skip()
    def OnGridRighClic(self,evt):
        self.PopupMenu(MyContextGrid(self), evt.GetPosition())
        evt.Skip()
    def CutData(self, evt):
        self.Copy()
        self.Delete()
        self.hasChanged= True
        self.hasSaved=   False
        evt.Skip()

    def CopyData(self, evt):
        self.Copy()

    #def onPaste(self, evt):
    #    box= evt.box
    #    top, left, rows, cols= box
    #    maxcol= self.NumberCols
    #    # determinando el rango maximo para pegar las columnas
    #    rangoCols= min([left+cols, maxcol-1])
    #    for row in range(top, top+rows):
    #        for col in range(left, rangoCols):
    #            self.__cellChanged(row,col)
    #    evt.Skip()
    def RangeSelected(self, evt):
        if evt.Selecting():
            self.tl = evt.GetTopLeftCoords()
            self.br = evt.GetBottomRightCoords()
    # @Busy(__('Pasting data'))
    #def PasteData(self, evt, *args, **params):
    #    self.OnPaste()
    #    self.hasChanged= True
    #    self.hasSaved=   False
    #    evt.Skip()

    def DeleteCurrentCol(self, evt):
        currentRow, currentCol, rows,cols = self.GetSelectionBox()[0]
        if cols < 1:
            return
        # A wxpython bug was detected
        # the app crash when trying to delete a colum
        # with a custom attr
        # deleting all data
        if 1:
            self.clearCol(currentCol)
            self.SetColLabelValue(currentCol, self.generateLabel( currentCol))
        else:
            self.DeleteCols( pos = currentCol, numCols = 1)
        self.AdjustScrollbars( )
        self.hasChanged= True
        self.hasSaved=   False
        evt.Skip()

    def DeleteCurrentRow(self, evt):
        currentRow, currentCol, rows,cols = self.GetSelectionBox()[0]
        if rows < 1:
            return
        self.DeleteRows(currentRow, 1)
        self.AdjustScrollbars()
        self.hasChanged= True
        self.hasSaved=   False
        evt.Skip()
    @Busy(__('Selecting all cells'))
    def SelectAllCells(self, evt):
        self.SelectAll()
    def LoadFile(self, evt, **params):
        if not params.has_key('fullpath'):
            wildcard=  __("Suported files")+" (*.txt;*.csv;*.xls;*.xlsx;*.db;*.dbf;*.mdb)|*.txt;*.csv;*.xls;*.xlsx;*db;*.dbf*.mdb|"\
                "Excel Files (*.xlsx;*.xlsm;*.xls)|*.xlsx;*.xlsm;*.xls|"\
                "Access mdb (*.mdb)|*mdb|"\
                "Txt file (*.txt)|*.txt|"\
                "Csv file (*.csv)|*.csv|"\
                "Sqlite Database (*.db)|*.db|"\
                "dbase Files (*.dbf)|*.dbf"
            dlg = wx.FileDialog(self, __("Load Data File"), "","",
                                wildcard= wildcard,
                                style = wx.OPEN)
            try:
                icon = imageEmbed().logo16
                dlg.SetIcon(icon)
            except:
                pass

            if dlg.ShowModal() != wx.ID_OK:
                dlg.Destroy()
                return (False, None)

            fullPath= dlg.Path
        else:
            fullPath= params.pop('fullpath')
        # if the file is loaded then 
        if self.load( fullPath)[0]:
            self.hasChanged= True
            self.hasSaved= True
            # emptying the undo - redo buffer
            self.emptyTheBuffer( )
            self.path2file= fullPath
            if evt != None:
                evt.Skip()
            return ( True, os.path.split( fullPath)[-1])
        else:
            return ( False, None)
    def onSave(self, evt, *args, **params):
        try:
            return self.save(*args, **params)
        finally:
            evt.Skip()
    def onSaveAs(self, evt, *args, **params):
        dlg= wx.FileDialog(self, __("Save Data File"), "" , "",\
                       "SEI (*.xlsx)|*.xlsx| \
                        All Files (*.*)| *.*",
                       style= wx.SAVE|wx.OVERWRITE_PROMPT, )
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            if not path.endswith('.xlsx'):
                path = path+'.xlsx'
        else:
            evt.Skip()
            return
        self.path2file= path
        try:
            return self.save(self, *args, **params)
        finally:
            evt.Skip()
            
    def SaveXlsAs(self, evt):
        return self.SaveXls(None, True)
    ###############################
    # adds columns and rows to the grid
    def AddNCells(self, numcols, numrows, attr= None):
        if numcols > 0:
            coltypes = self.__columnTypes
            coltypes.extend([3]*numcols)
            self.__columnTypes = coltypes
        insert= self.AppendCols(numcols)
        insert= self.AppendRows(numrows)
        if attr != None:
            for colNumber in range(self.GetNumberCols() - numcols, self.GetNumberCols(), 1):
                #self.SetColLabelAlignment(wx.ALIGN_LEFT, wx.ALIGN_BOTTOM)
                self.SetColAttr( colNumber, attr)
        self.AdjustScrollbars()
        self.change= True

    # function finds out how many cols contain data - all in a list
    #(ColsUsed) which has col #'s
    def GetUsedCols(self):
        # improving the performance
        if not self.change:
            return self.usedCols

        ColsUsed = []
        colnums = []
        dat = ''
        #
        # remodeling the data for speeding up the code
        #
        #for col in range(self.GetNumberCols()):
        #    if self.GetCellValue(0, col) != '':
        #        ColsUsed.append(self.GetColLabelValue(col))
        #        colnums.append(col)

        #tmp = 0
        #for col in range(self.GetNumberCols()):
        #    for row in range(self.GetNumberRows()):
        #        dat = self.GetCellValue(row, col)
        #        if dat != '':
        #            tmp += 1
        #            break # it's just needed to search by the first element

        #    if tmp > 0:
        #        ColsUsed.append(self.GetColLabelValue(col))
        #        colnums.append(col)
        #        tmp = 0

        # listing all columns of the grid
        for col in range(self.GetNumberCols()):
            ColsUsed.append(self.GetColLabelValue(col))
            colnums.append(col)

        self.usedCols= (ColsUsed, colnums)
        self.hasChanged= False ### to be fixed
        return self.usedCols

    def save(self, path= None, *args, **params):
        if not isinstance(path, (str, unicode)):
            path = None
        if path == None:
            if self.path2file == None:
                dlg= wx.FileDialog(self, __("Save Data File"), "" , "",\
                               "SEI (*.xlsx)|*.xlsx| \
                                All Files (*.*)| *.*",
                               style= wx.SAVE|wx.OVERWRITE_PROMPT, )
                if dlg.ShowModal() == wx.ID_OK:
                    path = dlg.GetPath()
                    if not path.endswith('.xlsx'):
                        path = path+'.xlsx'
                # check if the file exist
            else:
                path = self.path2file
        if path == None:
            raise StandardError( __("The input path isn't correct")+":%s"%path)
        else:
            self.path2file= path

        #if os.path.exists( self.path2file):
        #    wb= load_workbook( self.path2file)#, optimized_write = True)
        #else:
        wb= Workbook( )#optimized_write = True)
        wb.title= self.name

        # writing the contents from the wx.sheet into the wb
        sh= wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for colNumber in range( len( self.GetUsedCols())):
            colValues= self.GetCol( colNumber)
            for row, value in enumerate( colValues):
                sh.cell( row =row, column = colNumber).value =value
        wb.save( self.path2file)
        self.isSave= True
        return True

    def SaveXls(self, *args):
        if len(args) == 1:
            saveAs= False
        elif len(args) == 0:
            saveAs= True
        else:
            saveAs= args[1]
        self.reportObj= ReportaExcel(cell_overwrite_ok = True)
        if self.isSave == False or saveAs: # path del grid
            # mostrar el dialogo para guardar el archivo
            dlg= wx.FileDialog(self, __("Save Data File"), "" , "",\
                               "Excel (*.xlsx)|*.xlsx| \
                                Any (*.*)| *.*", wx.SAVE)
            if dlg.ShowModal() == wx.ID_OK:
                self.path = dlg.GetPath()
                if not self.path.endswith('.xls'):
                    self.path= self.path+'.xls'       
            else:
                return (False, None)
            self.path2file= self.reportObj.path = self.path
        else:
            self.path2file= self.reportObj.path = self.path
        cols, waste = self.GetUsedCols()
        if len(cols) == 0:
            pass
        else:
            rows = self.GetNumberRows()
            if rows == 0:
                print "Empty sheet"
                return
            totalResult = self.getByColumns(maxRow = rows)
            result= list()
            # reporting the header
            allColNames= [self.GetColLabelValue(col) for col in range(self.NumberCols) ]
            for posCol in range(waste[-1]+1):
                header= [  ]
                if posCol in waste:
                    header.append(allColNames[posCol])
                    header.extend(totalResult[posCol])
                else:
                    pass
                result.append(header)
            self.reportObj.writeByCols(result, sheet= 0) # report the data in the first sheet
        self.reportObj.save()
        self.isSave = True
        filename= os.path.split(self.path)[-1]
        if len( filename) > 8:
            filename = filename[:8]
        print "The file %s was successfully saved!" % self.reportObj.path
        return (True, filename)

    def load(self, path, *args, **params):
        # dispath load data depending on the file extension
        if path == None:
            return self._LoadFile(evt= None)
        
        extension= os.path.split(path)[1].split('.')[-1]
        
        available= {'csv':  self.LoadCsvTxt,
                    'txt':  self.LoadCsvTxt,
                    'xls':  self.LoadXls_xlsx,
                    'xlsx': self.LoadXls_xlsx,
                    'db':   self._LoadSqlite,
                    'dbf':  self._LoadDbf,
                    'mdb':  self._LoadMdb, ## load an sql database
                    'accdb':self._LoadMdb,
                    }
        try:
            result= available[extension](path, *args, **params)
            self.name= result[1]
            #self.path2file= path
            #if path.endswith('.xlsx'):
            #    self.name= result[1]
            #    self._wb= load_workbook(self.path2file)
            #    self._sh= self._wb.get_sheet_by_name(result[1])
            return result
        except KeyError:
            print 'Extension not available'

    def _LoadDbf(self, path, *arg, **params):
        from dbfpy import dbf
        db = dbf.Dbf(path)
        fieldNames= db.fieldNames
        fieldDefs= db.fieldDefs
        ##fieldDefs[colNumber].decodeValue(
        # if it's an empty database
        if len(fieldNames)== 0:
            return (True, path)

        for rowNumber, rec in enumerate(db):
            rowdata= [rec[fieldName] for colNumber, fieldName in enumerate(fieldNames)]
            self.PutRow(rowNumber, rowdata)

        # setting the col label values
        for colNumber, fieldName in enumerate(fieldNames):
            self.SetColLabelValue(colNumber, fieldName)
        return (True, path)

    def _LoadMdb(self, path, *args, **params):
        if path == None:
            return ( False, )
        
        import sys
        import os
        import adodbapi
        path2database= path
        ## http://social.msdn.microsoft.com/Forums/es-ES/0537e2e2-6c4d-42a0-ba77-b11a08893678/obtener-reportes-de-access
        path2SystemMdw= os.path.abspath(os.path.join(os.path.split(os.path.abspath(sys.argv[0]))[0],"System.mdw"))
        constr = "Provider=Microsoft.Jet.OLEDB.4.0;"+\
                 "Data Source=%s;"+\
                 "Jet OLEDB:System Database=" + path2SystemMdw
        constr= constr % path2database
        # connect to the database
        conn = adodbapi.connect(constr)
        # create a cursor
        try:
            cur = conn.cursor()
            cur.execute("SELECT * FROM MSysObjects WHERE Type = 1")
            result = cur.fetchall()
            # reading the existent tablenames
            tableNames= list([res for i,res in enumerate(result.ado_results) if 'MSysObjects' in res][0])
            allowTablenames2Import= [table for table in tableNames if not(table.startswith(u"MSys"))]
            # creating a dicctionay with all the columns of the existent tables
            tables= dict()
            for tablename in allowTablenames2Import:
                try:
                    # getting the columns of all tables
                    cur.execute("Select * from ["+ tablename +"]")
                    res= cur.fetchone()
                    tables[tablename]= res.rows.columnNames.keys()
                except:
                    continue
        finally:
            cur.close()
            conn.close()

        setting= {'Title': self.name,
                  '_size': wx.Size(220,300)}
        structure= list()
        # se pide la ruta de reporte se escribe una por defecto
        txt4 =     ('StaticText', (__('Select the table and the fields to import:'),))
        dbCtrl =  ('DataBaseImport', [tables,])
        structure.append([txt4])
        structure.append([dbCtrl])
        dlg= dialog(settings = setting, struct = structure)
        if dlg.ShowModal() == wx.ID_OK:
            values = dlg.GetValue()
            dlg.Destroy()
        else:
            dlg.Destroy()
            return
        tableName, fieldNames = values[0] #ordererString is missing
        ordererString = ''
        
        ## import the access data
        constr = "Provider=Microsoft.Jet.OLEDB.4.0;"+\
                 "Data Source=%s;" % path2database
        # connect to the database
        conn = adodbapi.connect(constr)
        simp= lambda x,y: x+","+"["+y+"]"
        # create a cursor
        try:
            cur = conn.cursor()
            # extract all the data
            #print  fieldNames
            names2extract= fieldNames
            if len(fieldNames)== 0:
                return
            names2extract[0]= "["+names2extract[0]+"]"
            sql = "select "+reduce( simp, names2extract) + " from "+ "["+ tableName+ "]"
            if ordererString in (None,''):
                sql += " " + ordererString
            names2extract[0]= names2extract[0][1:-1]
            #print sql
            cur.execute( sql)
            # show the result
            print __("conecting to the database")
            result = cur.fetchall()
            print __("conection sucsessfull")
            # close the cursor and connection
        finally:
            cur.close()
            conn.close()
        # sending the data to the report page
        # report.addRowData(names2extract, "consulta Base de datos", currRow= 0)
        #creating a new sheet in the Data Entry Panel
        tableName= self.name
        # self.grid.addRowData( [], tableName)
        requiredCols= len(result[0])-self.GetNumberCols()+1
        requiredRows= len(result)-self.GetNumberRows()+1
        if requiredCols > 0:
            self.AppendCols(requiredCols)
        if requiredRows > 0:
            self.AppendRows(requiredRows)
        # start the batch
        self.BeginBatch()
        try:
            for pos, row in enumerate(result):
                rowData= [getattr(row, prop) for prop in names2extract]
                newRowData= list()
                for rowi in rowData:
                    if rowi == None:
                        newRowData.append('')
                    else:
                        newRowData.append(rowi)
                self.Parent.Parent.addRowData( newRowData, currRow= pos) # need to be fixed because it dont neet to call the parent
            # labeling the columns
        finally:
            self.EndBatch()

        for pos, colNamei in enumerate(names2extract):
            self.SetColLabelValue(pos, colNamei)

        return ( True, path)

    def _LoadSqlite(self, path, *args, **params):
        # to load data from an sqlite database
        from sqlalchemy import create_engine
        if path == None:
            return ( False, )
        engine= create_engine( 'sqlite:///%s'%path, echo=False, )
        return ( self._loadDb( engine), path)
        
    def _loadDb(self, engine):
        from gridLib.gridsql import selectDbTableDialog, GenericDBClass
        dlg= selectDbTableDialog( self, engine, allow2edit= True)
        if dlg.ShowModal() == wx.ID_OK:
            values= dlg.GetValue()
        else:
            dlg.Destroy()
            return False
        # the dialog is destroyed after the results of the database
        value, filterTxt = values
        if value == None:
            dlg.Destroy()
            # The user didn't select any table
            return False
        # reading the data by columns and paste into the current sheet
        table=  dlg.m_grid.table
        sesion = table.Session()
        # add a page to write the result
        # self.addPage(name= 'noname', gridSize= (sesion.query(GenericDBClass).limit(20000).count(), len(table.colLabels)))
        for colNumber, colName in enumerate( table.colLabels):
            rowValues= list()
            for rowi in sesion.query( GenericDBClass).filter(filterTxt).limit(20000).all():
                rowValues.append( getattr( rowi, colName))
                # report the values
            # writing the data in a new sheet
            self.PutCol( colNumber, rowValues,)
        dlg.Destroy()
        return True
        
    def LoadCsvTxt(self, fullPath):
        from numpy import genfromtxt
        '''use the numpy library to load the data'''
        # comments='#', delimiter=None, skiprows=0, skip_header=0, skip_footer=0, converters=None, missing='', missing_values=None, filling_values=None, usecols=None, names=None, excludelist=None, deletechars=None, replace_space='_', autostrip=False, case_sensitive=True, defaultfmt='f%i', unpack=None, usemask=False, loose=True, invalid_raise=True
        btn1= ['FilePath',    [fullPath] ]
        txt1= ['StaticText',  [__('comments symbol')] ]
        btn2= ['TextCtrl',    [] ]
        txt2= ['StaticText',  [__('delimiter symbol')]]
        txt3= ['StaticText',  [__('Number of header lines to skip')]]
        btn3= ['IntTextCtrl', []]
        txt4= ['StaticText',  [__('Number of footer lines to skip')]]
        btn4= ['CheckBox',    [__('Has Header')]]

        structure= []
        structure.append([btn1 ])
        structure.append([btn2, txt1])
        structure.append([btn2, txt2])
        structure.append([btn3, txt3])
        structure.append([btn3, txt4])
        structure.append([btn4])

        setting = {'Title': __('Select a sheet')}

        dlg = dialog(self, struct= structure, settings= setting)
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return False, None

        (f_name, comments, delimiter, header2skip, footer2skip, hasHeader)= dlg.GetValue()
        if delimiter== u'':
            delimiter= None

        if header2skip == None:
            header2skip= 0

        if footer2skip == None:
            footer2skip= 0

        if comments == u'':
            comments= '#'

        dlg.Destroy()
        # reading the data
        data = genfromtxt(f_name, comments= comments,
                          dtype= None,
                          delimiter=  delimiter,
                          skip_header= header2skip,
                          skip_footer= footer2skip)
        # putting the data inito the Data Entry Panel
        if hasHeader:
            initRow= 1
        else:
            initRow= 0

        #grid= wx.GetApp().inputGrid 
        if len(data.shape)== 1:
            # it's considered to be of one column
            data.shape= (data.shape[0],1)
        for col in range(data.shape[1]):
            self.PutCol( col, data[initRow:,col])

        # Renaming the column of the Data Entry Panel
        if hasHeader:
            headerData= [data[0, col] for col in range(data.shape[1]) ]
            for pos,x in enumerate(headerData):
                if not isinstance( x, (str, unicode)):
                    x.__str__()
                # writing the data
                self.SetColLabelValue(pos, x)

        self.hasChanged= True
        self.hasSaved=   True
        return (True, os.path.split(f_name)[-1])

    def LoadXls_xlsx(self, fullPath):
        print 'import xlrd'
        filename= fullPath
        ##filenamestr= filename.__str__()
        #print '# remember to write an  r   before the path'
        print 'filename= r' + "'" + filename + "'"
        # se lee el libro
        wb= xlrd.open_workbook(filename)
        print 'wb = xlrd.open_workbook(filename)'
        sheets= [wb.sheet_by_index(i) for i in range(wb.nsheets)]
        print 'sheets= [wb.sheet_by_index(i) for i in range(wb.nsheets)]'
        sheetNames = [sheet.name for sheet in sheets]
        print 'sheetNames= ' + sheetNames.__str__()
        bt1= ('Choice',     [sheetNames])
        bt2= ('StaticText', [__('Select a sheet to be loaded')])
        bt3= ('CheckBox',   [__('Has header')])
        setting = {'Title': __('Select a sheet'),
                   '_size':  wx.Size(250,220)}

        dlg = dialog(self, struct=[[bt1,bt2],[bt3]], settings= setting)
        if dlg.ShowModal() != wx.ID_OK:
            return (False, None)

        (sheetNameSelected, hasHeader)= dlg.GetValue()
        if sheetNameSelected == None:
            return (False, None)

        if not isinstance(sheetNameSelected, (str, unicode)):
            sheetNameSelected = sheetNameSelected.__str__()

        print 'sheetNameSelected= ' + "'" + sheetNameSelected + "'"
        print 'hasHeader= ' + hasHeader.__str__()
        dlg.Destroy()

        if not ( sheetNameSelected in sheetNames):
            return (False, None)

        self._loadXls(sheetNameSelected= sheetNameSelected,
                      sheets= sheets,
                      sheetNames= sheetNames,
                      filename= filename,
                      hasHeader= hasHeader)
        print '''grid._loadXls(sheetNameSelected= sheetNameSelected,
                      sheets= sheets,
                      sheetNames= sheetNames,
                      filename= filename,
                      hasHeader= hasHeader)'''

        print __('Importing  : %s successful')%filename
        self.hasChanged= True
        self.hasSaved=   True
        return (True, sheetNameSelected)

    def _loadXls( self, *args,**params):
        sheets=         params.pop( 'sheets')
        sheetNames=     params.pop( 'sheetNames')
        sheetNameSelected= params.pop( 'sheetNameSelected')
        filename=       params.pop( 'filename')
        hasHeader=      params.pop( 'hasHeader')
        for sheet, sheetname in zip(sheets, sheetNames):
            if sheetname == sheetNameSelected:
                sheetSelected= sheet
                break

        #<p> updating the path related to the new open file
        self.path= filename
        self.hasSaved= True
        # /<p>

        # se hace el grid de tamanio 1 celda y se redimensiona luego
        self.ClearGrid()
        # reading the size of the needed sheet
        currentSize= (self.NumberRows-1, self.NumberCols-1)
        # se lee el tamanio de la pagina y se ajusta las dimensiones
        neededSize = (sheetSelected.nrows, sheetSelected.ncols)
        if neededSize[0]-currentSize[0] > 0:
            self.AppendRows(neededSize[0]-currentSize[0])

        if neededSize[1]-currentSize[1] > 0:
            self.AppendCols(neededSize[1]-currentSize[1])

        # se escribe los datos en el grid
        try:
            DECIMAL_POINT= wx.GetApp().DECIMAL_POINT
        except AttributeError:
            pass
        star= 0
        if hasHeader:
            star= 1
            for col in range( neededSize[1]):
                header= sheetSelected.cell_value( 0, col)
                if header == u'' or header == None:
                    ## return the column to it's normal label value
                    self.SetColLabelValue( col, self.generateLabel( col))
                elif not isinstance( header,( str, unicode)):
                    self.SetColLabelValue( col, sheetSelected.cell_value( 0, col).__str__())
                else:
                    self.SetColLabelValue( col, sheetSelected.cell_value( 0, col))
        else:
            # return all header to default normal value
            for col in range( neededSize[1]):
                self.SetColLabelValue(col, self.generateLabel( col))

        if hasHeader and neededSize[0] < 2:
            return

        book_datemode = sheetSelected.book.datemode

        for reportRow, row in enumerate(range( star, neededSize[0])):
            for col in range( neededSize[1]):
                newValue = sheetSelected.cell_value( row, col)
                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                if isinstance( newValue, (str, unicode)):
                    self.SetCellValue( reportRow, col, newValue)
                elif sheetSelected.cell_type( row, col) in ( 2,): # number
                    self.SetCellValue( reportRow, col, str( newValue).replace('.', DECIMAL_POINT))
                elif sheetSelected.cell_type( row, col) in ( 3,): # date
                    year, month, day, hour, minute, second = xlrd.xldate_as_tuple(newValue, book_datemode)
                    self.SetCellValue( reportRow, col, self.datetime.date(year, month, day).__str__())
                else:
                    try:
                        self.SetCellValue (reportRow, col, str(newValue))
                    except:
                        print  __("Could not import the row,col (%i,%i)") % (row+1, col+1)

    def generateLabel( self, colNumber):
        colNumber += 1
        analyse = True
        result = list()
        while analyse:
            res = colNumber/26.0
            if res == int(res):
                result.append(26)
                colNumber = colNumber/26-1
                analyse = res > 1
                continue
            fp = res-int(res) # float Part
            # deleting fix by rounding
            if fp !=0 :       fp = int(round(fp*26.0,0))
            else:             fp = 1
            result.append(fp)
            colNumber = colNumber/26
            analyse = res > 1
        res = '' 
        while len(result):
            res += chr(result.pop(-1) + 64) 
        return res

    def CleanRowData(self, row):
        indata = []
        missingvalue= wx.GetApp().missingvalue
        for i in range(self.GetNumberCols()):
            datapoint = self.GetCellValue(row, i)
            if (datapoint != ''):
                value = float(datapoint)
                if (value != missingvalue):
                    indata.append(value)
        return indata
    # Routine to return a "clean" list of data from one column
    def CleanData(self, coldata):
        indata = []
        self.missing = 0
        try:
            dp= wx.GetApp().DECIMAL_POINT
        except AttributeError:
            dp= DECIMAL_POINT
        missingvalue= wx.GetApp().missingvalue 
        if dp == '.':
            for i in range(self.GetNumberRows()):
                datapoint = self.GetCellValue(i, coldata).strip()
                if (datapoint != u'' or datapoint.replace(' ','') != u'') and (datapoint != u'.'):
                    try:
                        value = float(datapoint)
                        if (value != missingvalue):
                            indata.append(value)
                        else:
                            self.missing = self.missing + 1
                    except ValueError:
                        pass
        else:
            for i in range(self.GetNumberRows()):
                datapoint = self.GetCellValue(i, coldata).strip().replace(dp, '.')
                if (datapoint != u'' or datapoint.replace(' ','') != u'') and (datapoint != u'.'):
                    try:
                        value = float(datapoint)
                        if (value != missingvalue):
                            indata.append(value)
                        else:
                            self.missing = self.missing + 1
                    except ValueError:
                        pass
        return indata

    def _cleanData(self, data, hasHeader= False):
        if isinstance(data, (str, unicode)):
            data= [data]

        if not isiterable(data):
            raise TypeError('Only iterable data allowed!')

        for pos in range(len(data)-1, -1, -1):
            if data[pos] != u'':
                pos= pos+1
                break

        data= data[:pos + [0, 1][hasHeader]]
        # changing data into a numerical value
        try:
            dp = wx.GetApp().DECIMAL_POINT
        except AttributeError:
            dp= DECIMAL_POINT
        result= list()
        for dat in data:
            if dat == u'' or dat.replace(' ','') == u'': # to detect a cell of only space bar
                dat = None
            else:
                try:
                    dat= float(dat.replace(dp, '.'))
                except:
                    pass

            result.append(dat)
        return result

    @__controlColnumber
    def GetCol(self, col, hasHeader= False):
        return self._cleanData( self._getCol( col, hasHeader))

    def PutRow(self, rowNumber, data):
        try: 
            if isinstance(rowNumber, (str, unicode)):
                if not(rowNumber in self.rowNames):
                    raise TypeError(__('You can only use a numeric value, or the name of an existing row'))
                for pos, value in enumerate(self.rowNames):
                    if value == rowNumber:
                        rowNumber= pos
                        break

            if not isnumeric(rowNumber):
                raise TypeError(__('You can only use a numeric value, or the name of an existing row'))

            rowNumber= int(rowNumber)        
            if rowNumber < 0 or rowNumber > self.GetNumberRows():
                raise StandardError(__('The minimum accepted col is 0, and the maximum is %i')%self.GetNumberRows()-1)

            self.clearRow(rowNumber)

            if isinstance( data,(str, unicode)):
                data= [data]

            if isinstance( data, (int, long, float)):
                data= [data]

            if isinstance( data, (ndarray),):
                data= ravel( data)

            cols2add= len( data) - self.GetNumberCols()
            if cols2add > 0:
                if len( data) > 16384:
                    data= data[:16384]
                    cols2add= len( data) - self.GetNumberCols()
                self.AppendCols( cols2add) ############### TO TEST

            try:
                dp= wx.GetApp().DECIMAL_POINT
            except AttributeError:
                dp= DECIMAL_POINT

            newdat= list()
            for row, dat in enumerate( data):
                if isinstance( dat, (str, unicode)):
                    try:
                        dat= str(float(dat.replace(dp,'.'))).replace('.',dp)
                    except:
                        pass
                else:
                    try:
                        dat= str(dat)
                    except:
                        dat= None

                newdat.append(dat)

            for col, dat in enumerate(newdat):
                try:
                    self.SetCellValue(rowNumber, col, dat)
                except UnicodeDecodeError:
                    self.SetCellValue(rowNumber, col, dat.decode("utf8","replace"))
                self.__cellChanged(rowNumber, col)
        except:
            raise
        finally:
            self.hasSaved= False
            self.hasChanged= True

    def GetRow(self, row):
        return self._cleanData( self._getRow( row))

    def GetEntireDataSet(self, numcols):
        """Returns the data specified by a list 'numcols' in a Numpy
        array"""
        import numpy.array
        biglist = []
        for i in range(len(numcols)):
            smalllist = wx.GetApp().frame.CleanData(numcols[i])
            biglist.append(smalllist)
        return numpy.array((biglist), numpy.float)

    @__controlColnumber
    def GetColNumeric(self, colNumber):
        # return only the numeric values of a selected colNumber or col label
        # all else values are drop
        # add the ability to manage non numerical values
        values= self._cleanData( self._getCol( colNumber))
        return [val for val in values if not isinstance(val,(unicode, str)) and val != None ]

    @__controlColnumber
    def _getCol(self, colNumber, includeHeader= False):
        return self._getColNumber(colNumber, includeHeader)

    @__controlColnumber
    def _getColNumber(self, colNumber, includeHeader= False):
        result = [self.GetCellValue(row, colNumber) for row in range(self.GetNumberRows())]
        if includeHeader:
            result.insert(0, self.GetColLabelValue(colNumber))
        return result

    def _getRow( self, rowNumber):
        if isinstance( rowNumber, (str, unicode)):
            # searching for a col with the name:
            if not(rowNumber in self.rowNames):
                raise TypeError('You can only use a numeric value, or the name of an existing row')
            for pos, value in enumerate(self.rowNames):
                if value == rowNumber:
                    rowNumber= pos
                    break

        if not isnumeric(rowNumber):
            raise TypeError('You can only use a numeric value, or the name of an existing row')

        if rowNumber > self.GetNumberRows():
            raise StandardError('The maximum row allowed is %i, but you selected %i'%(self.GetNumberRows()-1, rowNumber))

        return self._getRowNumber(rowNumber)

    def _getRowNumber(self, rowNumber):
        if not isnumeric( rowNumber):
            raise TypeError('Only allow numeric values for the row, but you input '+ type(rowNumber).__str__())

        rowNumber= int(rowNumber)
        if rowNumber < 0 or rowNumber > self.GetNumberRows():
            raise StandardError('The minimum accepted row is 0, and the maximum is %i'%self.GetNumberRows()-1)

        return [self.GetCellValue( rowNumber, col) for col in range( self.GetNumberCols())]
    
    def PutCol(self, *args, **params):
        return self.putCol(*args, **params)

    @__controlColnumber
    def putCol( self, colNumber, data):
        try:
            self.clearCol( colNumber)
            if isinstance( data,(str, unicode)):
                data= [data]
            if isinstance( data, (int, long, float)):
                data= [data]
            if isinstance( data, (ndarray),):
                data= ravel( data)
            rows2add= len( data) - self.GetNumberRows()+1
            if rows2add > 0:
                if len( data) > 1e6:
                    data= data[:1e6]
                    rows2add= len( data) - self.GetNumberRows()
                self.AppendRows( rows2add)
            try:
                dp= wx.GetApp().DECIMAL_POINT
            except:
                dp= DECIMAL_POINT
            newdat= list()
            for row, dat in enumerate( data):
                if isinstance( dat, (str, unicode)):
                    try:
                        dat= str(float(dat.replace(dp,'.'))).replace('.',dp)
                    except:
                        pass
                else:
                    try:
                        dat= str(dat)
                    except:
                        dat= None
                newdat.append(dat)
            for row, dat in enumerate(newdat):
                try:
                    self.SetCellValue(row, colNumber, dat)
                except UnicodeDecodeError:
                    self.SetCellValue(row, colNumber, dat.decode("utf8","replace"))
        except:
            raise
        finally:
            self.change= True

    @__controlColnumber
    def clearCol( self, colNumber):
        for row in range( self.GetNumberRows()):
            self.SetCellValue( row, colNumber, u'')
    @__controlColnumber
    def setColType(self, colNumber, colType, *args):
        """Setting a type for an existent column
        :param colNumber: number or name of the column
        :param colType: type of the column as integer
        colType as an string"""
        if isinstance(colType,(str, unicode) ):
            colType= colType.upper()
            if not COLTYPES.has_key(colType):
                if not (colType in COLTYPES.values()):
                    colType= 3
                else:
                    for key, value in COLTYPES.items():
                        if value == colType:
                            break
                colType= key
            colType = int( colType)
        if not colType in range( 5):
            raise StandardError( __("Column type not allowed"))

        attr = wx.grid.GridCellAttr()
        if colType == 0:
            renderer = GE.DATE()
            attr.SetEditor(renderer)
        elif colType == 1:
            listData= []
            # check the previus control
            ###print self.__columnTypes[colNumber]
            if  self.__columnTypes[colNumber] == 1: # LIST type
                # retreiving the data of the control
                currEditor= self.GetCellEditor(0,colNumber)
                try:
                    ctrl = currEditor.GetControl()
                    listData = ctrl.GetItems()
                except:
                    pass

            if len(args) == 0:
                # show a dialog to the user to select its options
                editableList = Ctrl.EditableListBox(__("Input the available options"), listData)
                structure = list()
                structure.append([editableList])
                dlg= dialog(struct= structure)
                if dlg.ShowModal() == wx.ID_OK:
                    values=dlg.GetValue()
                    dlg.Destroy()
                else:
                    dlg.Destroy()
                    return
                options = values[0]
            else:
                # if the user pass the options by a command
                options= args[0]
            #"allowOthers" allows the user to create new selection items on the fly.
            editor = wx.grid.GridCellChoiceEditor(options, allowOthers = False)
            attr.SetEditor(editor)
        elif colType == 2: # Float
            editor = wx.grid.GridCellFloatEditor()
            attr.SetEditor(editor)
        elif colType == 3: # variant
            pass
        elif colType == 4: # Integer
            editor =  wx.grid.GridCellNumberEditor()
            attr.SetEditor(editor)
        self.SetColAttr(colNumber, attr)
        self.__columnTypes[colNumber] = colType
        
    def clearRow( self, rowNumber):
        if rowNumber < 0 or rowNumber > self.GetNumberRows():
            raise StandardError( 'The minimum accepted row is 0, and the maximum is %i'%self.GetNumberRols()-1)

        for col in range( self.GetNumberCols()):
            self.SetCellValue( rowNumber, col, u'')
            
    def get_selection(self):
        """ Returns an index list of all cell that are selected in 
        the grid. All selection types are considered equal. If no 
        cells are selected, the current cell is returned.
        
        from: http://trac.wxwidgets.org/ticket/9473"""

        dimx, dimy = (self.NumberRows, self.NumberCols) #self.parent.dimensions[:2]
        selection = []
        selection += self.GetSelectedCells()
        selected_rows = self.GetSelectedRows()
        selected_cols = self.GetSelectedCols()
        selection += list((row, y) for row in selected_rows for y in xrange(dimy))
        selection += list((x, col) for col in selected_cols for x in xrange(dimx)) 
        for tl,br in zip(self.GetSelectionBlockTopLeft(), self.GetSelectionBlockBottomRight()):
            selection += [(x,y) for x in xrange(tl[0],br[0]+1) for y in xrange(tl[1], br[1]+1)]
            if selection == []: 
                selection = self.get_currentcell()
        selection = sorted(list(set(selection)))
        return selection
        
    #def setColNames(self,names):
    #    # escribe los nombres de las columnas en el grid
    #    if not(type(names) == type(list()) or type(names) == type(tuple())):
    #        raise TypeError("It's allowed one list")
    #    [self.SetColLabelValue(colNumber, value) for colNumber, value in enumerate(names) ]
        
    def setRowNames(self,names):
        if not(type(names) == type(list()) or type(names) == type(tuple())):
            raise TypeError("It's allowed one iterable list")
        [self.SetRowLabelValue(rowNumber, value) for rowNumber, value in enumerate(names)]
        
    def getByColumns(self, maxRow = None):
        # retorna el valor de la malla por columnas
        numRows = self.GetNumberRows()
        ncols= self.GetNumberCols()
        if maxRow != None:
            numRows= min([numRows, maxRow])
        # se extrae los contenidos de cada fila
        return tuple([ self.GetCol(col)[:numRows] ] for col in range(ncols))

    def getValue(self):
        # retorma los contenidos de la malla ordenados por filas y 
        # empezando por el encabezado
        # se extrae el nombre de la columnas
        numCols = self.GetNumberCols()
        contenidoGrid = [self.colNames]
        contenidoGrid.extend([tuple([self.GetCellValue(row,col) for col in range(numCols)]) for row in range(self.GetNumberRows())])
        return tuple(contenidoGrid)
    
def test():
    # para verificar el funcionamiento correcto del grid
    pass

if __name__ == '__main__':    
        app = wx.PySimpleApp()
        app.translate= lambda x:x
        frame = wx.Frame(None, -1, size=(700,500), title = "wx example")
        grid = NewGrid(frame)
        grid.CreateGrid(2000,60)
        grid.SetDefaultColSize(70, 1)
        grid.EnableDragGridSize(False)
        grid.SetCellValue(0,0,"Col is")
        grid.SetCellValue(1,0,"Read ")
        grid.SetCellValue(1,1,"hello")
        grid.SetCellValue(2,1,"23")
        grid.SetCellValue(4,3,"greren")
        grid.SetCellValue(5,3,"geeges")
        # make column 1 multiline, autowrap
        cattr = wxCellAttr()
        cattr.SetEditor(wxCellAutoWrapStringEditor())
        #cattr.SetRenderer(wxCellAutoWrapStringRenderer())
        grid.SetColAttr(1, cattr)
        frame.Show(True)
        app.MainLoop()
